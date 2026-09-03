#!/usr/bin/env python3
"""Generate the public suggested-resources page from the vault's signed register.

Spec SITE-06 D4. Reads `Suggested-Resources.md`, writes the `resources` page node
into `src/content/site-content.json`, and writes `scripts/resource-link-hosts.json`
for `cdt00-origin-scan.mjs` to union into its known-link-hosts set.

Follows `build_evaluation_form.py`'s shape: digest the source, refuse while
unsigned, print what it wrote.

## The round-trip is proven before anything is written

SITE-06 finding 11. `src/content/site-content.json` round-trips through Python at
`indent=2, ensure_ascii=False` with a trailing newline, and at no other setting;
`indent=4` and `ensure_ascii=True` each fail. Nothing in the repo enforces that,
and SITE-02 and SITE-05 both write to the same file. So this script re-serialises
what it read and refuses unless the bytes are identical, BEFORE it edits. A
concurrent reformat stops this generator rather than being amplified by it.

## The floor

Two or more items publishes an area as a `linkGrid`. Fewer publishes a one-line
stub saying it is still being gathered. **No area is ever omitted**, because the
six areas are visible to anyone who read the workshop materials and a reader
cannot tell an omission from a judgment.

Usage
-----
    python3 scripts/build_resources_page.py --print-items          # parse only
    python3 scripts/build_resources_page.py --check                # all gates, no write
    python3 scripts/build_resources_page.py --apply \
        --allow-unsigned-register --reason "shipping before the 4 Sep send-out"
"""

import argparse
import hashlib
import json
import os
import re
import sys
from urllib.parse import urlparse

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
VAULT = os.path.expanduser(
    "~/Documents/Josh & Katie Vault/Claude Can Access PARA")
TRACK = os.path.join(VAULT, "Projects/OBT/OBT Consultant Track")

REGISTER = os.path.join(TRACK, "Suggested-Resources.md")
KSA = os.path.join(TRACK, "Psalms (Bali 2026)/OBT CDT Psalms KSAs.md")
CONTENT = os.path.join(REPO, "src/content/site-content.json")
HOSTS = os.path.join(REPO, "scripts/resource-link-hosts.json")
RULES = os.path.join(REPO, "scripts/resource-content-rules.json")

PAGE_ID = "resources"
PAGE_ROUTE = "/suggested-resources"

# D1's cap on the assembled small-caps line. `note` renders at 0.75rem in
# uppercase with wide tracking, so a value long enough to wrap twice reads as a
# mistake rather than as information. Measured against the rendered card in
# criterion 7 rather than guessed.
NOTE_CAP = 150

FIELDS = ("title", "what", "where", "cost", "url", "from")


def die(message: str) -> None:
    sys.exit(f"REFUSED: {message}")


def slug(text: str) -> str:
    out = re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")
    return re.sub(r"-{2,}", "-", out)[:60]


def read_front_matter(text: str) -> tuple[dict, str]:
    """Minimal YAML for the shapes this file actually uses.

    Deliberately not PyYAML: it is not a dependency of this repo, and the front
    matter here is a fixed shape written by one script and one person. A parser
    that accepts more than the shape would accept a typo as data.
    """
    if not text.startswith("---\n"):
        die(f"{REGISTER} has no YAML front matter.")
    end = text.index("\n---\n", 3)
    raw, body = text[4:end], text[end + 5:]

    data: dict = {}
    key = None
    current_list = None
    current_map = None
    for line in raw.split("\n"):
        if not line.strip() or line.strip().startswith("#"):
            continue
        indent = len(line) - len(line.lstrip())
        stripped = line.strip()

        if indent == 0 and ":" in stripped and not stripped.startswith("-"):
            key, _, value = stripped.partition(":")
            key, value = key.strip(), value.strip()
            current_list, current_map = None, None
            if value in ("", "[]"):
                data[key] = [] if value == "[]" else ""
                if value == "":
                    data[key] = ""
                    current_list = key
            else:
                data[key] = value.strip("'\"")
        elif stripped.startswith("- ") and key:
            item = stripped[2:].strip()
            if data.get(key) in ("", None) or isinstance(data.get(key), str):
                data[key] = []
            if ":" in item:
                k, _, v = item.partition(":")
                current_map = {k.strip(): v.strip().strip("'\"")}
                data[key].append(current_map)
            else:
                data[key].append(item.strip("'\""))
                current_map = None
        elif current_map is not None and ":" in stripped:
            k, _, v = stripped.partition(":")
            current_map[k.strip()] = v.strip().strip("'\"")
    return data, body


def parse_items(body: str) -> list[dict]:
    """One record per `**Item N**` block under an `## Area N. <title>` heading."""
    areas: list[dict] = []
    current = None
    item = None

    for line in body.split("\n"):
        heading = re.match(r"^##\s+Area\s+\d+\.\s+(.+?)\s*$", line)
        if heading:
            current = {"title": heading.group(1), "items": []}
            areas.append(current)
            item = None
            continue
        if re.match(r"^##\s+", line):
            # Any other H2 (the notes and reference sections) ends the areas.
            current, item = None, None
            continue
        if current is None:
            continue
        if re.match(r"^\*\*Item\s+\d+\*\*\s*$", line.strip()):
            item = {k: "" for k in FIELDS}
            current["items"].append(item)
            continue
        field = re.match(r"^-\s+(\w+):\s*(.*)$", line)
        if field and item is not None and field.group(1) in FIELDS:
            item[field.group(1)] = field.group(2).strip()

    # An item block with no title at all is an empty template slot, not an item.
    for area in areas:
        area["items"] = [i for i in area["items"] if i["title"]]
    return areas


def workshop_area_titles() -> list[str]:
    """The six areas, from the workshop's own KSA table.

    SITE-06 finding 9 and program finding 9: the six are the workshop's, not a
    list invented for a website. An area in the register that no workshop source
    names is refusal 5, which catches a renamed area rather than publishing both
    spellings.
    """
    with open(KSA, encoding="utf-8") as handle:
        text = handle.read()
    return re.findall(r"^\|\s*\*\*(.+?)\*\*\s*\|", text, re.M)


def load_rules() -> dict:
    with open(RULES, encoding="utf-8") as handle:
        return json.load(handle)


def roster_names(xlsx: str | None) -> list[str]:
    """Participant names, READ and never copied into this repo.

    SITE-06 finding 13. The roster grep the first draft proposed returned zero on
    `dist/` before this page existed, so it could not tell a correct page from an
    absent one. The gate that works is on the generator's INPUT, because the
    input is prose typed by people who do not know the rule.
    """
    if not xlsx:
        return []
    try:
        from openpyxl import load_workbook
    except ImportError:
        die("openpyxl is not installed, so the roster gate cannot run. "
            "Install it, or pass --no-roster-gate and say why in the build record.")
    workbook = load_workbook(xlsx, read_only=True, data_only=True)
    names: list[str] = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(c).strip() if c is not None else "" for c in rows[0]]
        first = headers.index("First Name") if "First Name" in headers else None
        last = headers.index("Family Name") if "Family Name" in headers else None
        for row in rows[1:]:
            parts = []
            for index in (first, last):
                if index is not None and index < len(row) and row[index]:
                    parts.append(str(row[index]).strip())
            if len(parts) == 2 and all(len(p) > 2 for p in parts):
                names.append(" ".join(parts))
    return sorted(set(names))


def walk_ids(node, out: list[str]) -> None:
    """`walkNode`'s exact traversal: the node, then `blocks`, then `items`.

    Reimplemented rather than imported. `nodeIndex()` is a TypeScript export and
    no `.mjs` or `.py` script in this repo can call it, and nothing here adds a
    loader for one. So this states its own traversal as the source of its count,
    and the count is printed rather than compared against a number that would
    match under a naive traversal too.
    """
    if isinstance(node, dict):
        if isinstance(node.get("id"), str):
            out.append(node["id"])
        for key in ("blocks", "items"):
            for child in node.get(key) or []:
                walk_ids(child, out)


def existing_ids(content: dict) -> list[str]:
    out: list[str] = []
    for root in [content["site"], *content["pages"], *content["workshops"]]:
        walk_ids(root, out)
    return out


def build_nodes(areas: list[dict], gathered_on: str) -> tuple[dict, list[str]]:
    """The page node, and every id it introduces."""
    published = [a for a in areas if len(a["items"]) >= 2]
    stubs = [a for a in areas if len(a["items"]) < 2]

    intro = (
        "What to read next, in the six areas the Psalms workshop itself is built on. "
        "Every item says where to get it and what it costs, because a recommendation "
        "you cannot find or cannot afford is not one.\n\n"
        "These six are what one workshop taught. The Five Threads are what a consultant "
        "is formed in across the whole track, so a thread runs through several of these "
        "areas rather than matching one.\n\n"
        f"Last gathered {gathered_on}. The list grows as the people who teach these areas add to it."
    )

    blocks: list[dict] = [{
        "id": f"{PAGE_ID}.00.intro",
        "type": "prose",
        "body": intro,
    }]

    for area in areas:
        key = slug(area["title"])
        if len(area["items"]) >= 2:
            items = []
            for item in area["items"]:
                note = " · ".join(p for p in (item["where"], item["cost"], item["from"]) if p)
                node = {
                    "id": f"{PAGE_ID}.{key}.{slug(item['title'])}",
                    "type": "cta",
                    "label": item["title"],
                    "body": item["what"],
                    "note": note,
                }
                if item["url"]:
                    node["href"] = item["url"]
                items.append(node)
            blocks.append({
                "id": f"{PAGE_ID}.{key}",
                "type": "linkGrid",
                "title": area["title"],
                "items": items,
            })
        else:
            have = len(area["items"])
            blocks.append({
                "id": f"{PAGE_ID}.{key}",
                "type": "prose",
                "title": area["title"],
                "body": (
                    f"Still being gathered. {'One suggestion is in' if have else 'Nothing is in'} "
                    "so far, and an area publishes once it has two. If you know what belongs here, "
                    "write to Joshua."
                ),
            })

    page = {
        "id": PAGE_ID,
        "route": PAGE_ROUTE,
        "navLabel": "Reading",
        "title": "What to read next",
        "metaDescription": (
            "Suggested reading for oral Bible translation consultants, in the six areas "
            "the OBT Consultant Development Track's Psalms workshop is built on."
        ),
        "kicker": "Suggested resources",
        "blocks": blocks,
    }

    ids: list[str] = []
    walk_ids(page, ids)
    return page, ids, published, stubs


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--apply", action="store_true", help="write; omit for a dry run")
    parser.add_argument("--check", action="store_true", help="run every gate and write nothing")
    parser.add_argument("--print-items", action="store_true", help="parse and print, then stop")
    parser.add_argument("--allow-unsigned-register", action="store_true")
    parser.add_argument("--reason", help="required with --allow-unsigned-register")
    parser.add_argument("--roster-xlsx", help="workbook to read participant names from, for refusal 7")
    parser.add_argument("--no-roster-gate", action="store_true",
                        help="skip refusal 7. Say why in the build record.")
    parser.add_argument("--indent", type=int, default=2, help="for the round-trip mutation only")
    parser.add_argument("--register", help="read a different register file. For mutation tests, "
                                           "so a gate can be watched going red without editing "
                                           "the real contract document.")
    args = parser.parse_args()

    global REGISTER
    if args.register:
        REGISTER = os.path.abspath(args.register)
        print(f"!! reading a substitute register: {REGISTER}")

    with open(REGISTER, encoding="utf-8") as handle:
        source = handle.read()
    digest = hashlib.sha256(source.encode()).hexdigest()[:16]
    front, body = read_front_matter(source)
    areas = parse_items(body)

    print(f"register digest {digest}")
    print(f"{len(areas)} area(s) parsed, {sum(len(a['items']) for a in areas)} item(s)")
    for area in areas:
        mark = "publishes" if len(area["items"]) >= 2 else "STUB"
        print(f"  {len(area['items']):>2}  {mark:<9}  {area['title']}")

    if args.print_items:
        for area in areas:
            print(f"\n## {area['title']}")
            for item in area["items"]:
                print(f"  - {item['title']}")
                for field in FIELDS[1:]:
                    print(f"      {field}: {item[field] or '(empty)'}")
        return

    # ---- Refusal 1: unsigned ------------------------------------------------
    signed = str(front.get("signed_off", "false")).lower() == "true"
    if not signed:
        if not args.allow_unsigned_register:
            die("`signed_off` is false in the register's front matter. "
                "Sign it, or pass --allow-unsigned-register --reason \"...\".")
        if not args.reason:
            die("--allow-unsigned-register needs --reason. "
                "This register is regenerated for years by whoever is nearest, and an "
                "override with no recorded reason becomes the normal way to run it.")
        print(f"\n!! running against an UNSIGNED register")
        print(f"!! reason: {args.reason}")

    # ---- Refusal 5: an area no workshop source names ------------------------
    canonical = workshop_area_titles()
    print(f"\n{len(canonical)} area title(s) in the workshop's KSA table")
    unknown = [a["title"] for a in areas if a["title"] not in canonical]
    if unknown:
        die("area(s) in the register that no workshop source names:\n  " +
            "\n  ".join(unknown) +
            "\nThe workshop's own titles are:\n  " + "\n  ".join(canonical))
    missing = [t for t in canonical if t not in [a["title"] for a in areas]]
    if missing:
        die("the register omits area(s) the workshop names:\n  " + "\n  ".join(missing) +
            "\nNo area may be omitted; an area below the floor publishes as a stub.")
    print("  ok: the register's areas and the workshop's are the same set")

    # ---- Refusals 2, 3, 6: per item -----------------------------------------
    consented = {
        str(c.get("name", "")).strip()
        for c in (front.get("contributors") or [])
        if isinstance(c, dict) and str(c.get("named", "")).lower() == "yes"
    }
    print(f"\ncontributors with a yes consent flag: {len(consented)}"
          f"{' (' + ', '.join(sorted(consented)) + ')' if consented else ' — asserted empty, not assumed'}")

    for area in areas:
        for item in area["items"]:
            where = item["where"]
            if item["url"] and not where:
                die(f"item {item['title']!r} in {area['title']!r} has a url and no `where`. "
                    "A link is not an answer to where a person gets a book.")
            if not where:
                die(f"item {item['title']!r} in {area['title']!r} has no `where`.")
            if not item["cost"]:
                die(f"item {item['title']!r} in {area['title']!r} has no `cost`. "
                    "Cost is a fact and never a comment: `Free`, a figure, or the programme it comes with.")
            if item["from"] and item["from"] not in consented:
                die(f"item {item['title']!r} names {item['from']!r} in `from`, and that "
                    "contributor has no yes consent flag in the front matter.")
            note = " · ".join(p for p in (where, item["cost"], item["from"]) if p)
            if len(note) > NOTE_CAP:
                die(f"item {item['title']!r} assembles a {len(note)}-character small-caps line, "
                    f"over the {NOTE_CAP} cap:\n  {note}\n"
                    "Shorten `where` or `cost`; it renders in uppercase at 0.75rem and wraps badly.")
    print("  ok: every item has `where` and `cost`, no unconsented name, no over-long note line")

    # ---- Refusals 7 and 8: content rules over the whole document ------------
    rules = load_rules()
    lowered = source.lower()

    def scan(label: str, phrases: list[str]) -> list[str]:
        return [p for p in phrases if re.search(r"\b" + re.escape(p.lower()) + r"\b", lowered)]

    if args.no_roster_gate:
        print("\n!! refusal 7's roster half is SKIPPED (--no-roster-gate)")
        roster = []
    else:
        roster = roster_names(args.roster_xlsx)
        if not roster:
            die("the roster gate has an empty population, so it cannot fail and would "
                "report success over nothing. Pass --roster-xlsx <workbook>, or "
                "--no-roster-gate and say why in the build record.")
    hits = scan("roster", roster)
    if hits:
        die(f"participant name(s) from the roster appear in the register: {hits}. "
            "No participant name reaches a public page.")
    if roster:
        print(f"\nrefusal 7, roster half: {len(roster)} name(s) checked, 0 found")

    for group, label in (("partner_orgs", "partner organization"),
                         ("pathway_contrast", "pathway-contrast phrase"),
                         ("do_not_route", "do-not-route programme")):
        phrases = rules[group]["phrases"]
        hits = scan(group, phrases)
        if hits:
            die(f"{label}(s) named in the register: {hits}. See scripts/resource-content-rules.json "
                f"for why this is a refusal and not a warning.")
        print(f"refusal {'7' if group != 'do_not_route' else '8'}, {label}: "
              f"{len(phrases)} phrase(s) checked, 0 found")

    # ---- The round-trip, proven before anything is written ------------------
    with open(CONTENT, "rb") as handle:
        original_bytes = handle.read()
    content = json.loads(original_bytes.decode())
    reserialised = (json.dumps(content, indent=args.indent, ensure_ascii=False) + "\n").encode()
    print(f"\nround-trip at indent={args.indent}, ensure_ascii=False: "
          f"{len(original_bytes)} bytes read, {len(reserialised)} bytes re-serialised")
    if reserialised != original_bytes:
        die(f"site-content.json does not round-trip at indent={args.indent}. "
            "It round-trips at indent=2, ensure_ascii=False and at no other setting "
            "(SITE-06 finding 11). Either the setting is wrong, or another spec has "
            "reformatted the file and this generator must not amplify that.")
    print("  ok: the file round-trips, so this script may edit it")

    # ---- Refusal 4: a duplicate node id -------------------------------------
    before = existing_ids(content)
    page, new_ids, published, stubs = build_nodes(areas, str(front.get("gathered_on", "")))
    print(f"\n{len(before)} node id(s) in site-content.json today, "
          f"{len(set(before))} distinct")
    if len(before) != len(set(before)):
        dupes = sorted({i for i in before if before.count(i) > 1})
        die(f"site-content.json already carries duplicate id(s) before this run: {dupes}. "
            "A later node silently replaces an earlier one in nodeIndex().")

    existing_page = next((p for p in content["pages"] if p["id"] == PAGE_ID), None)
    owned = set()
    if existing_page:
        walk_ids(existing_page, (owned_list := []))
        owned = set(owned_list)
    # Two directions, and the first was missing until a mutation test found it.
    # `set(new_ids) & existing` collapses an internal duplicate before it is
    # compared, so two items with the same title in one area produced two
    # identical ids and this gate reported no clash. The self-check has to come
    # first, because a page that shadows its own node never reaches the second
    # comparison at all.
    if len(new_ids) != len(set(new_ids)):
        internal = sorted({i for i in new_ids if new_ids.count(i) > 1})
        die(f"this page generates duplicate id(s) within itself: {internal}. "
            "Two items with the same title in one area slug to one id, and the second "
            "silently replaces the first in nodeIndex(). Retitle one of them.")

    clash = sorted(set(new_ids) & (set(before) - owned))
    if clash:
        die(f"node id(s) already used elsewhere in site-content.json: {clash}. "
            "A duplicate id does not error; the later node replaces the earlier one in "
            "nodeIndex(), so Joshua's edit-in-place silently changes the wrong node.")
    print(f"{len(new_ids)} new id(s), all distinct, 0 clash with the "
          f"{len(set(before) - owned)} id(s) this page does not own")

    print(f"\n{len(published)} area(s) publish as a linkGrid, {len(stubs)} as a stub")
    for area in stubs:
        print(f"  stub: {area['title']}")

    hosts = sorted({urlparse(i["url"]).hostname for a in areas for i in a["items"] if i["url"]})
    print(f"\n{len(hosts)} distinct link host(s):")
    for host in hosts:
        print(f"  {host}")

    if args.check or not args.apply:
        print("\nEvery gate passed. Nothing written (use --apply).")
        return

    pages = [p for p in content["pages"] if p["id"] != PAGE_ID]
    # Keep the page next to the other public pages and before the member pages,
    # so nav order stays readable in the file itself.
    insert_at = next((i for i, p in enumerate(pages) if p["id"] == "members"), len(pages))
    pages.insert(insert_at, page)
    content["pages"] = pages

    with open(CONTENT, "w", encoding="utf-8") as handle:
        handle.write(json.dumps(content, indent=2, ensure_ascii=False) + "\n")

    with open(HOSTS, "w", encoding="utf-8") as handle:
        json.dump({
            "_readme": (
                "Written by build_resources_page.py. cdt00-origin-scan.mjs unions these "
                "into KNOWN_LINK_HOSTS so a new outbound LINK host on the resources page "
                "does not print as NEW. A link is navigation, not contact, and only a "
                "contacted origin fails that scan."
            ),
            "generated_from": os.path.basename(REGISTER),
            "source_digest": digest,
            "hosts": hosts,
        }, handle, indent=2)
        handle.write("\n")

    print(f"\nwrote the `{PAGE_ID}` page node to src/content/site-content.json")
    print(f"wrote {len(hosts)} host(s) to scripts/resource-link-hosts.json")
    print(f"source_digest {digest}")


if __name__ == "__main__":
    main()
