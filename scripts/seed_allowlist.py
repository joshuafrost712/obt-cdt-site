#!/usr/bin/env python3
"""Seed `member_allowlist` from the OBT-CDT sign-up workbook.

The allowlist decides one thing only: who may create a portal account at all.
It is NOT an authorization boundary for content — RLS is. So the cost of
including someone who never attends is a dormant row, while the cost of
excluding someone who did attend is a person who cannot reach their own report
and has to email Joshua. The default therefore takes the UNION of every valid
address anywhere in the workbook, not just the Master tab.

**No participant address is ever written into this repository.** The workbook
stays outside it (export it with rclone, see the runbook below), addresses are
masked in output unless you ask otherwise, and the only place they land is the
portal database.

Usage
-----
    # 1. export the sheet somewhere outside the repo
    rclone backend copyid gdrive: <SHEET_ID> /tmp/signup.xlsx

    # 2. look at what would be written (default; writes nothing)
    python3 scripts/seed_allowlist.py --xlsx /tmp/signup.xlsx

    # 3. write it
    set -a; . ~/.claude/secrets/obt-cdt-portal.env; set +a
    python3 scripts/seed_allowlist.py --xlsx /tmp/signup.xlsx --apply

Environment
-----------
    PORTAL_URL         https://<project-ref>.supabase.co
    PORTAL_SECRET_KEY  the `sb_secret_...` key (service role). Read from the
                       environment, never a command-line argument, because
                       argv is visible to every process on the machine and
                       lands in shell history.

Requires `openpyxl` (already present in this machine's Homebrew python3).
"""

import argparse
import json
import os
import re
import sys
import urllib.error
import urllib.request
from collections import defaultdict

# Deliberately permissive on the local part and strict about the shape. The
# workbook has cells holding two addresses, trailing commas, and stray text, so
# this scans for addresses rather than validating a whole cell.
EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")

# Which columns actually name a participant. Anything not listed is ignored, so
# a facilitator column or a free-text note cannot quietly enroll someone.
# (tab, column header) -> short label recorded in `member_allowlist.note`.
SOURCES = {
    ("Master", "Primary Email"): "master",
    ("Master", "Alternate emails"): "master-alt",
    ("Psalms 2026", "Email"): "psalms-2026",
    ("Hebrew Training", "Email"): "hebrew-training",
    ("Sheet1", "Running list of eligible OBT Workshop Attendees"): "eligible-legacy",
    ("Sheet1", "Attendees at Chiang Mai"): "chiang-mai-2025",
}


def normalize(address: str) -> str:
    """Match `handle_new_portal_user`'s rule exactly: lower(btrim(email)).

    If these two ever disagree, a person on the list is refused at signup with
    a message that says they are not on it. Keep them identical.
    """
    return address.strip().lower()


def mask(address: str) -> str:
    """First and last character of the local part.

    Keeping the last character matters: two colleagues at one organization
    often share a local-part length, and a mask that hides both ends renders
    them as the same string on screen — which makes a review of this output
    silently useless at exactly the moment it should catch something.
    """
    local, _, domain = address.partition("@")
    if len(local) <= 2:
        return f"{local[:1] or '?'}*@{domain}"
    return f"{local[0]}{'*' * (len(local) - 2)}{local[-1]}@{domain}"


def near_duplicates(addresses) -> list[tuple[str, ...]]:
    """Addresses that are probably one person.

    This is the failure that actually bites: a report published to
    `x@partner.br` when the profile registered `x@partner.com.br` does not error,
    it lands in the unmatched queue looking like a stranger. Catching it here,
    before anyone signs up, is far cheaper than reconciling it afterwards — and
    each pair found is a candidate row for `member_alias`.
    """
    groups: dict[str, set[str]] = defaultdict(set)
    for address in addresses:
        local, _, domain = address.partition("@")
        # Gmail-style dots and +tags are noise; so is a co./com. domain shift.
        key_local = local.split("+")[0].replace(".", "")
        root = ".".join(p for p in domain.split(".") if p not in {"com", "co"})
        groups[f"{key_local}@{root}"].add(address)
    return [tuple(sorted(v)) for v in groups.values() if len(v) > 1]


def harvest(path: str) -> dict[str, set[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl is not installed: python3 -m pip install openpyxl")

    workbook = load_workbook(path, read_only=True, data_only=True)
    found: dict[str, set[str]] = defaultdict(set)
    seen_tabs = set()

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(c).strip() if c is not None else "" for c in rows[0]]
        for index, header in enumerate(headers):
            label = SOURCES.get((sheet.title, header))
            if not label:
                continue
            seen_tabs.add((sheet.title, header))
            for row in rows[1:]:
                if index >= len(row) or row[index] is None:
                    continue
                for match in EMAIL_RE.finditer(str(row[index])):
                    found[normalize(match.group(0))].add(label)

    missing = set(SOURCES) - seen_tabs
    if missing:
        # Loudly, not silently. A renamed tab would otherwise drop a whole
        # cohort off the allowlist and the run would still say "done".
        for tab, column in sorted(missing):
            print(f"  !! expected column not found: [{tab}] {column!r}", file=sys.stderr)
        sys.exit("Refusing to seed from a workbook whose shape has changed.")

    return found


def request(method: str, url: str, key: str, body=None, prefer=None):
    data = json.dumps(body).encode() if body is not None else None
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
    }
    if prefer:
        headers["Prefer"] = prefer
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=30) as response:
            raw = response.read().decode()
            return json.loads(raw) if raw.strip() else []
    except urllib.error.HTTPError as error:
        sys.exit(f"{method} {url} -> {error.code}: {error.read().decode()[:500]}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", required=True, help="path to the exported workbook (keep it outside this repo)")
    parser.add_argument("--apply", action="store_true", help="actually write; omit for a dry run")
    parser.add_argument("--show", action="store_true", help="print full addresses instead of masking them")
    args = parser.parse_args()

    harvested = harvest(args.xlsx)
    if not harvested:
        sys.exit("No addresses found. Wrong file?")

    by_source: dict[str, int] = defaultdict(int)
    for labels in harvested.values():
        for label in labels:
            by_source[label] += 1

    print(f"{len(harvested)} distinct addresses across {len(by_source)} sources")
    for label in sorted(by_source, key=lambda k: -by_source[k]):
        print(f"  {by_source[label]:>3}  {label}")

    rows = [
        {"email": address, "note": ",".join(sorted(labels))}
        for address, labels in sorted(harvested.items())
    ]

    show = (lambda a: a) if args.show else mask

    pairs = near_duplicates(harvested)
    if pairs:
        print(f"\n{len(pairs)} possible duplicate person(s) — check before anyone signs up:")
        for pair in pairs:
            print("  " + "  ≈  ".join(show(a) for a in pair))
        print("  Both are allowlisted either way. If they are one person, the")
        print("  second address belongs in `member_alias`, or a report sent to it")
        print("  will sit unmatched looking like a stranger.")

    if not args.apply:
        print("\nDry run. Would upsert:")
        for row in rows:
            print(f"  {show(row['email']):<40} {row['note']}")
        print("\nRe-run with --apply to write.")
        return

    url = os.environ.get("PORTAL_URL", "").rstrip("/")
    key = os.environ.get("PORTAL_SECRET_KEY", "")
    if not url or not key:
        sys.exit("Set PORTAL_URL and PORTAL_SECRET_KEY (see the module docstring).")

    before = request("GET", f"{url}/rest/v1/member_allowlist?select=email", key)
    existing = {r["email"] for r in before}

    # merge-duplicates so a re-run is boring: the note is refreshed, added_at is
    # left alone by the table default, and nobody is removed. Removal is
    # deliberate work, not a side effect of re-running an import.
    request(
        "POST",
        f"{url}/rest/v1/member_allowlist",
        key,
        body=rows,
        prefer="resolution=merge-duplicates,return=minimal",
    )

    after = request("GET", f"{url}/rest/v1/member_allowlist?select=email", key)
    added = {r["email"] for r in after} - existing
    print(f"\nUpserted {len(rows)}. {len(added)} new, {len(rows) - len(added)} already present.")
    print(f"Allowlist now holds {len(after)}.")

    orphans = existing - {r["email"] for r in rows}
    if orphans:
        # Not deleted. Someone may have been added by hand for a good reason.
        print(f"\n{len(orphans)} address(es) in the allowlist are absent from this workbook:")
        for address in sorted(orphans):
            print(f"  {show(address)}")


if __name__ == "__main__":
    main()
