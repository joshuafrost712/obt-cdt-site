#!/usr/bin/env python3
"""Seed the 41-unit CBC competency registry from the vault sources.

Spec CDT-01 D2. The point of this script is not that it inserts rows. It is that
**the registry and its sources cannot silently disagree**: it refuses to run
against sources whose shape has changed, it records a digest of everything it
read, re-running it against unchanged sources is a no-op, and a change to an
existing unit's identity-bearing text is refused rather than applied.

Posture copied from scripts/seed_allowlist.py, which is this repo's model for a
script that touches real data: credentials from the environment and never from a
command line, dry-run by default, masked output, and a refusal when the source's
shape has changed.

Usage
-----
    # look, change nothing (the default)
    python3 scripts/seed_competency_registry.py

    # write
    export SUPABASE_URL=https://<ref>.supabase.co
    export SUPABASE_SECRET_KEY=sb_secret_...        # service role; never logged
    python3 scripts/seed_competency_registry.py --apply

    # a competency statement legitimately changed upstream
    python3 scripts/seed_competency_registry.py --apply --allow-unit-change

    # emit the CIT self-assessment workbook
    python3 scripts/seed_competency_registry.py --emit-sheet /path/outside/repo.xlsx

Exit codes: 0 fine, 1 a refusal you should read, 2 a usage error.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import urllib.error
import urllib.request
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from lib.cbc_matrix import (  # noqa: E402
    EXPECTED,
    MatrixError,
    Registry,
    display_counts,
    load,
)

# The vault is not inside this repo and never will be: it holds participant data.
# Overridable so a scratch copy can be gated in a test.
DEFAULT_VAULT = Path(
    os.environ.get(
        "OBT_CDT_VAULT",
        str(Path.home() / "Documents/Josh & Katie Vault/Claude Can Access PARA"),
    )
)
MATRIX_SUBDIR = "Projects/cbc-competency/cbc-matrix"
DOMAIN_MAP_SUBDIR = "Projects/OBT/OBT Consultant Track/Intake Assessment/Instruments/Domain-Map.md"

# The seven data tables, in insert order. Dependencies first: a unit references a
# category, a descriptor references a unit.
TABLE_ORDER = [
    "competency_scale",
    "competency_domain",
    "competency_category",
    "category_domain",
    "competency_unit",
    "unit_descriptor",
    "unit_prerequisite",
]

PREREQUISITES = [
    # The one seeded row. Unit 28 is Biblical Languages; level 1 is Awareness.
    {"unit_key": "U28", "min_level": 1, "gates": "Entry to OBT-CDT Hebrew workshops"},
]


def mask(value: str) -> str:
    if not value:
        return "(unset)"
    return value[:8] + "…" + f"({len(value)} chars)"


class Rest:
    """The smallest possible Supabase REST client. No dependency."""

    def __init__(self, url: str, key: str):
        self.url = url.rstrip("/")
        self.key = key

    def _req(self, method: str, path: str, body=None, prefer: str | None = None):
        req = urllib.request.Request(
            f"{self.url}/rest/v1/{path}", method=method
        )
        req.add_header("apikey", self.key)
        req.add_header("Authorization", f"Bearer {self.key}")
        req.add_header("Content-Type", "application/json")
        if prefer:
            req.add_header("Prefer", prefer)
        data = json.dumps(body).encode() if body is not None else None
        try:
            with urllib.request.urlopen(req, data, timeout=60) as r:
                raw = r.read().decode()
                return json.loads(raw) if raw.strip() else []
        except urllib.error.HTTPError as e:
            detail = e.read().decode()[:400]
            raise SystemExit(
                f"\n{method} {path} failed with HTTP {e.code}.\n  {detail}\n"
                "  A 401 or 403 here usually means SUPABASE_SECRET_KEY is a\n"
                "  publishable key rather than a service-role key. These tables are\n"
                "  revoked from every client role by design (CDT-00 question 7)."
            )

    def select(self, table: str, columns: str = "*") -> list[dict]:
        return self._req("GET", f"{table}?select={columns}")

    def upsert(self, table: str, rows: list[dict], on_conflict: str) -> None:
        if not rows:
            return
        for i in range(0, len(rows), 200):
            self._req(
                "POST",
                f"{table}?on_conflict={on_conflict}",
                rows[i : i + 200],
                prefer="resolution=merge-duplicates,return=minimal",
            )

    def insert(self, table: str, rows: list[dict]) -> list[dict]:
        if not rows:
            return []
        return self._req("POST", table, rows, prefer="return=representation")


def rows_for(reg: Registry) -> dict[str, list[dict]]:
    return {
        "competency_scale": [
            {"level": l, "label": lab, "definition": d} for l, lab, d in reg.scale
        ],
        "competency_domain": [
            {"domain_key": d["domain_key"], "name": d["name"], "ordinal": d["ordinal"]}
            for d in reg.domains
        ],
        "competency_category": [
            {
                "category_key": c["category_key"],
                "track": c["track"],
                "name": c["name"],
                "ordinal": c["ordinal"],
            }
            for c in reg.categories
        ],
        "category_domain": [
            {
                "category_key": l["category_key"],
                "domain_key": l["domain_key"],
                "is_primary": l["is_primary"],
                "note": l["note"],
            }
            for l in reg.links
        ],
        "competency_unit": [
            {
                "unit_key": u.unit_key,
                "category_key": u.category_key,
                "sub_area": u.sub_area,
                "statement": u.statement,
                "rationale": u.rationale,
                "ordinal": u.ordinal,
            }
            for u in reg.units
        ],
        "unit_descriptor": [
            {"unit_key": u.unit_key, "ordinal": i, "text": t}
            for u in reg.units
            for i, t in enumerate(u.descriptors, start=1)
        ],
        "unit_prerequisite": PREREQUISITES,
    }


def report_counts(reg: Registry) -> None:
    print("counts, re-measured from the sources in this run")
    for key, expected in EXPECTED.items():
        got = reg.counts[key]
        print(f"  {key:<20} {got:>4}   expected {expected}")
    primary, displayed = display_counts(reg)
    print(f"  units over PRIMARY links  {sum(primary.values()):>3}   (the invariant, must be 41)")
    print(f"  units over ALL links      {sum(displayed.values()):>3}   (the display set, 42 on purpose)")
    print(f"  per domain, primary  : {primary}")
    print(f"  per domain, displayed: {displayed}")
    print(f"  source digest: {reg.source_digest[:16]}…")
    print(
        f"  Domain-Map.md signed off: "
        f"{'yes' if reg.domain_map_signed_off else 'NO (see --allow-unsigned-domain-map)'}"
    )


def diff_units(existing: list[dict], desired: list[dict]) -> tuple[list, list, list]:
    """New units, identity-bearing changes, and units that would disappear."""
    by_key = {r["unit_key"]: r for r in existing}
    new, changed = [], []
    for d in desired:
        cur = by_key.get(d["unit_key"])
        if cur is None:
            new.append(d)
            continue
        for f in ("statement", "category_key", "ordinal"):
            if str(cur.get(f)) != str(d[f]):
                changed.append(
                    {
                        "unit_key": d["unit_key"],
                        "field": f,
                        "old_value": str(cur.get(f)),
                        "new_value": str(d[f]),
                    }
                )
    gone = [r["unit_key"] for r in existing if r["unit_key"] not in {d["unit_key"] for d in desired}]
    return new, changed, gone


def emit_sheet(reg: Registry, path: Path) -> None:
    """The CIT self-assessment workbook. Every unit key is machine-written.

    Columns match self_assessment_intake, NOT the ledger: the ledger is CDT-03's
    and does not exist yet, and CDT-03's drain owns the rename.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        raise SystemExit(
            "--emit-sheet needs openpyxl:  python3 -m pip install openpyxl"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Self-assessment"
    headers = [
        "unit_key",
        "Category",
        "Sub-area",
        "Competency statement",
        "claim_status",
        "claimed_level",
        "note",
    ]
    ws.append(headers)
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDE7F0")
    for c in ws[1]:
        c.font = bold
        c.fill = fill
    for u in reg.units:
        ws.append(
            [u.unit_key, u.category_name, u.sub_area or "", u.statement, "", None, ""]
        )

    dv_status = DataValidation(
        type="list", formula1='"claimed,no-claim,skipped"', allow_blank=True
    )
    dv_level = DataValidation(type="whole", operator="between", formula1=0, formula2=3, allow_blank=True)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_level)
    dv_status.add(f"E2:E{len(reg.units) + 1}")
    dv_level.add(f"F2:F{len(reg.units) + 1}")

    widths = [10, 26, 24, 78, 14, 14, 34]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    for row in ws.iter_rows(min_row=2, max_col=4):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    # The four scale points, so a CIT is not guessing what 2 means.
    gs = wb.create_sheet("Rating scale")
    gs.append(["level", "label", "definition"])
    for c in gs[1]:
        c.font = bold
    for l, lab, d in reg.scale:
        gs.append([l, lab, d])
    gs.column_dimensions["A"].width = 8
    gs.column_dimensions["B"].width = 20
    gs.column_dimensions["C"].width = 100
    for row in gs.iter_rows(min_row=2, max_col=3):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    note = wb.create_sheet("How to fill this in")
    for line in [
        ["A blank row is recorded as no-claim: considered, claiming nothing. It is NOT"],
        ["a zero. If you have not looked at a unit yet, write skipped, which says so."],
        ["claim_status = claimed  means you are claiming the level in claimed_level."],
        ["claim_status = no-claim means you have considered it and claim nothing."],
        ["claim_status = skipped  means you are deliberately passing over it for now."],
        [""],
        ["A claim of 0 is a real claim: put claimed in claim_status and 0 in claimed_level."],
        ["An empty claimed_level with claim_status = claimed will be refused on import."],
        [""],
        ["Do not edit the unit_key column. It is generated, and the importer refuses"],
        ["the whole file if a key does not match the registry."],
    ]:
        note.append(line)
    note.column_dimensions["A"].width = 95

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"wrote {path}  ({len(reg.units)} units, keys generated)")


def _sql_literal(v) -> str:
    """A Postgres literal for the handful of types these row dicts hold."""
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return repr(v)
    if isinstance(v, (list, tuple)):
        inner = ", ".join(_sql_literal(x) for x in v)
        return f"array[{inner}]"
    return "'" + str(v).replace("'", "''") + "'"


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("--apply", action="store_true", help="write; default is dry-run")
    ap.add_argument(
        "--emit-sql",
        type=Path,
        help="write the rows as SQL to this path instead of a database, for a "
        "test harness that seeds them inside a rolled-back transaction",
    )
    ap.add_argument(
        "--allow-unit-change",
        action="store_true",
        help="permit a change to an existing unit's statement, category or ordinal",
    )
    ap.add_argument(
        "--allow-unsigned-domain-map",
        action="store_true",
        help="write even though Domain-Map.md is not signed off",
    )
    ap.add_argument("--vault", type=Path, default=DEFAULT_VAULT)
    ap.add_argument("--matrix-dir", type=Path, default=None, help="override, for tests")
    ap.add_argument("--domain-map", type=Path, default=None, help="override, for tests")
    ap.add_argument("--emit-sheet", type=Path, default=None)
    ap.add_argument("--show", action="store_true", help="unmask credentials in output")
    args = ap.parse_args()

    matrix_dir = args.matrix_dir or (args.vault / MATRIX_SUBDIR)
    domain_map = args.domain_map or (args.vault / DOMAIN_MAP_SUBDIR)

    print(f"matrix     {matrix_dir}")
    print(f"domain map {domain_map}\n")

    try:
        reg = load(matrix_dir, domain_map)
    except MatrixError as e:
        print("REFUSED, and this is the script working as intended:\n", file=sys.stderr)
        print(f"  {e}\n", file=sys.stderr)
        return 1

    report_counts(reg)

    if args.emit_sheet:
        print()
        emit_sheet(reg, args.emit_sheet)
        return 0

    desired = rows_for(reg)
    total = sum(len(v) for v in desired.values())
    print(f"\n{total} rows across {len(TABLE_ORDER)} tables")
    for t in TABLE_ORDER:
        print(f"  {t:<22} {len(desired[t]):>4}")

    if args.emit_sql:
        # The same rows this script would upsert, as SQL, so a test harness can
        # seed them inside a transaction it will roll back. See the identical flag
        # on seed_bundles.py for why this exists rather than a second parser.
        # Runs after the gate, so a source set that fails the gate emits nothing.
        lines = [
            "-- Generated by scripts/seed_competency_registry.py --emit-sql. Do not edit.",
            "-- Derived from an UNSIGNED Domain-Map. For use inside a transaction that",
            "-- will be rolled back. Never commit them.",
            "",
        ]
        for t in TABLE_ORDER:
            if not desired[t]:
                continue
            cols = list(desired[t][0].keys())
            lines.append(f"-- {t}: {len(desired[t])} rows")
            lines.append(f"insert into public.{t} ({', '.join(cols)}) values")
            lines.append(
                ",\n".join(
                    "  (" + ", ".join(_sql_literal(r[c]) for c in cols) + ")"
                    for r in desired[t]
                )
                + ";"
            )
            lines.append("")
        args.emit_sql.write_text("\n".join(lines), encoding="utf-8")
        print(f"\nemitted {total} rows as SQL to {args.emit_sql}")
        print("nothing written to any database, and no credentials were read.")
        return 0

    if not args.apply:
        print(
            "\ndry run, nothing written. Re-run with --apply once SUPABASE_URL and\n"
            "SUPABASE_SECRET_KEY are set. The counts above are the whole gate; if it\n"
            "printed them, the sources are in the shape this registry expects."
        )
        return 0

    url = os.environ.get("SUPABASE_URL", "")
    key = os.environ.get("SUPABASE_SECRET_KEY", "") or os.environ.get(
        "SUPABASE_SERVICE_ROLE_KEY", ""
    )
    if not url or not key:
        print(
            "\n--apply needs SUPABASE_URL and SUPABASE_SECRET_KEY in the environment.\n"
            "  Never pass a key as a command-line argument: it lands in shell history\n"
            "  and in the process list.",
            file=sys.stderr,
        )
        return 2
    print(f"\nurl {url}")
    print(f"key {key if args.show else mask(key)}")

    if not reg.domain_map_signed_off and not args.allow_unsigned_domain_map:
        print(
            f"\nREFUSED: {domain_map.name} is not signed off.\n"
            "  The macro-domain grouping is an instrument-track judgment, reviewed by\n"
            "  Viji, and five of the 26 assignments are called arguable. Writing it as\n"
            "  fact before that review would make a draft look decided.\n"
            "  Set `signed_off: true` in that file, or pass --allow-unsigned-domain-map\n"
            "  if you are deliberately seeding a staging database.",
            file=sys.stderr,
        )
        return 1

    rest = Rest(url, key)

    existing_units = rest.select("competency_unit", "unit_key,category_key,statement,ordinal")
    new_units, changed, gone = diff_units(existing_units, desired["competency_unit"])
    print(f"\nunits: {len(existing_units)} present, {len(new_units)} new, "
          f"{len(changed)} identity-bearing change(s), {len(gone)} would disappear")
    for c in changed:
        print(f"  {c['unit_key']}.{c['field']}")
        print(f"     old {c['old_value'][:90]!r}")
        print(f"     new {c['new_value'][:90]!r}")
    if gone:
        print(f"  units in the database but not in the sources: {gone}")

    if changed and not args.allow_unit_change:
        print(
            "\nREFUSED: an existing unit's identity-bearing text changed.\n"
            "  Every rating ever recorded points at a unit_key. Rewriting the\n"
            "  statement under that key silently reassigns all of them, and the\n"
            "  counts stay 41/26/194 so no gate catches it.\n"
            "  If CBC really did revise this unit, re-run with --allow-unit-change\n"
            "  and the change is written to unit_revision with the old text kept.",
            file=sys.stderr,
        )
        return 1

    prior = rest.select("registry_version", "version,source_digest")
    if prior and any(p["source_digest"] == reg.source_digest for p in prior):
        print(
            "\nno change: a registry_version already records this exact source digest.\n"
            "  Nothing written, no registry_version row added."
        )
        return 0

    conflicts = {
        "competency_scale": "level",
        "competency_domain": "domain_key",
        "competency_category": "category_key",
        "category_domain": "category_key,domain_key",
        "competency_unit": "unit_key",
        "unit_descriptor": "unit_key,ordinal",
        "unit_prerequisite": "unit_key,gates",
    }
    for t in TABLE_ORDER:
        rest.upsert(t, desired[t], conflicts[t])
        print(f"  wrote {t:<22} {len(desired[t]):>4}")

    version_row = rest.insert(
        "registry_version",
        [
            {
                "source_digest": reg.source_digest,
                "unit_count": reg.counts["units"],
                "category_count": reg.counts["categories"],
                "descriptor_count": reg.counts["descriptor_bullets"],
                "domain_count": reg.counts["domains"],
                "link_count": reg.counts["category_links"],
                "note": "seed_competency_registry.py"
                + ("; --allow-unit-change" if args.allow_unit_change else "")
                + ("; unsigned domain map" if not reg.domain_map_signed_off else ""),
            }
        ],
    )
    version = version_row[0]["version"] if version_row else None
    print(f"  registry_version {version}")

    if changed:
        rest.insert(
            "unit_revision",
            [{**c, "registry_version": version} for c in changed],
        )
        print(f"  unit_revision    {len(changed)} row(s) recording the change")

    print("\ndone.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
