#!/usr/bin/env python3
"""Load a filled CIT self-assessment workbook into self_assessment_intake.

Spec CDT-01 D3. This is the manual fallback that wave 1 actually runs on: the
first roughly eight assessment sessions happen before the portal has a URL, so
the sheet is the real intake path and not a courtesy.

Two rules it exists to enforce.

**It refuses the WHOLE file on any unknown unit_key.** Not "imports the rows that
matched". A partial import of a competency ledger is worse than no import,
because the missing rows read as "no claim", and "no claim" is a meaningful state
that an evaluator will act on. A silent gap here becomes a wrong conversation
about someone's competence.

**An empty cell is not a zero.** `claim_status` carries the distinction:
`no-claim` with a null level is "I considered this and claim nothing", while
`claimed` with level 0 is "I claim none of this competency". Those are different
statements about a person and the schema keeps them different.

No participant address ever enters this repo. The workbook is passed by path and
lives outside it, exactly as scripts/seed_allowlist.py requires.

Usage
-----
    # validate only, against the registry, writing nothing (the default)
    python3 scripts/import_self_assessment.py --sheet /path/outside/repo/filled.xlsx

    # validate against the vault sources rather than the database, for when the
    # project is not reachable yet
    python3 scripts/import_self_assessment.py --sheet f.xlsx --registry-from-vault

    # write
    export SUPABASE_URL=https://<ref>.supabase.co
    export SUPABASE_SECRET_KEY=sb_secret_...
    python3 scripts/import_self_assessment.py --sheet f.xlsx --email person@example.org --apply

Exit codes: 0 fine, 1 a refusal you should read, 2 a usage error.
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from lib.cbc_matrix import MatrixError, load  # noqa: E402
from seed_competency_registry import (  # noqa: E402
    DEFAULT_VAULT,
    DOMAIN_MAP_SUBDIR,
    MATRIX_SUBDIR,
    Rest,
    mask,
)

VALID_STATUS = {"claimed", "no-claim", "skipped"}
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[A-Za-z]{2,}$")


def read_sheet(path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("needs openpyxl:  python3 -m pip install openpyxl")
    if not path.exists():
        raise SystemExit(f"no such workbook: {path}")

    wb = load_workbook(path, data_only=True)
    if "Self-assessment" not in wb.sheetnames:
        raise SystemExit(
            f"{path.name}: no sheet named 'Self-assessment'. Sheets present: "
            f"{wb.sheetnames}. This importer reads the workbook that "
            "seed_competency_registry.py --emit-sheet produces."
        )
    ws = wb["Self-assessment"]
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    required = ["unit_key", "claim_status", "claimed_level"]
    missing = [c for c in required if c not in header]
    if missing:
        raise SystemExit(
            f"{path.name}: the header is missing {missing}. Found: {header}\n"
            "  Do not rename the generated columns; the importer keys on them."
        )
    idx = {name: i for i, name in enumerate(header)}
    rows = []
    for r, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if raw is None or all(v is None or str(v).strip() == "" for v in raw):
            continue
        rows.append(
            {
                "row": r,
                "unit_key": (str(raw[idx["unit_key"]]).strip() if raw[idx["unit_key"]] is not None else ""),
                "claim_status": (str(raw[idx["claim_status"]]).strip().lower() if raw[idx["claim_status"]] is not None else ""),
                "claimed_level": raw[idx["claimed_level"]],
                "note": (str(raw[idx["note"]]).strip() if "note" in idx and raw[idx["note"]] is not None else ""),
            }
        )
    return rows


def validate(rows: list[dict], known_keys: set[str]) -> list[dict]:
    """Return normalised rows, or raise with every problem found at once.

    Every problem, not the first: a person fixing a spreadsheet should get the
    whole list rather than discover them one run at a time.
    """
    problems: list[str] = []
    seen: dict[str, int] = {}
    out: list[dict] = []

    for r in rows:
        key = r["unit_key"]
        if key not in known_keys:
            problems.append(
                f"  row {r['row']}: unit_key {key!r} is not in the registry"
            )
            continue
        if key in seen:
            problems.append(
                f"  row {r['row']}: unit_key {key} already appeared on row {seen[key]}"
            )
            continue
        seen[key] = r["row"]

        status = r["claim_status"]
        level = r["claimed_level"]

        # An untouched row is a no-claim, which is a real answer and not an error.
        if status == "" and (level is None or str(level).strip() == ""):
            out.append({"unit_key": key, "claim_status": "no-claim", "claimed_level": None, "note": r["note"]})
            continue

        if status == "" and level is not None:
            problems.append(
                f"  row {r['row']}: a level of {level!r} with no claim_status. "
                "Write 'claimed' in claim_status, or clear the level."
            )
            continue
        if status not in VALID_STATUS:
            problems.append(
                f"  row {r['row']}: claim_status {status!r} is not one of "
                f"{sorted(VALID_STATUS)}"
            )
            continue

        if status == "claimed":
            if level is None or str(level).strip() == "":
                problems.append(
                    f"  row {r['row']}: claim_status is 'claimed' but claimed_level is "
                    "empty. A claim needs a level; use 'no-claim' if you claim nothing."
                )
                continue
            try:
                lv = int(float(str(level).strip()))
            except ValueError:
                problems.append(f"  row {r['row']}: claimed_level {level!r} is not a number")
                continue
            if lv < 0 or lv > 3:
                problems.append(
                    f"  row {r['row']}: claimed_level {lv} is outside 0 to 3"
                )
                continue
            out.append({"unit_key": key, "claim_status": "claimed", "claimed_level": lv, "note": r["note"]})
        else:
            if level is not None and str(level).strip() != "":
                problems.append(
                    f"  row {r['row']}: claim_status is {status!r} but claimed_level is "
                    f"{level!r}. Only a 'claimed' row carries a level."
                )
                continue
            out.append({"unit_key": key, "claim_status": status, "claimed_level": None, "note": r["note"]})

    if problems:
        raise MatrixError(
            f"refusing the whole file: {len(problems)} problem(s).\n"
            + "\n".join(problems)
            + "\n\n  Nothing was imported. A partial import is worse than none here, "
            "because\n  the rows that did not land read as 'no claim', and that is a "
            "meaningful\n  state an evaluator would act on."
        )
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("--sheet", type=Path, required=True, help="path OUTSIDE this repo")
    ap.add_argument("--email", help="the subject's address; required with --apply")
    ap.add_argument("--apply", action="store_true")
    ap.add_argument(
        "--registry-from-vault",
        action="store_true",
        help="validate keys against the vault sources instead of the database",
    )
    ap.add_argument("--vault", type=Path, default=DEFAULT_VAULT)
    ap.add_argument("--show", action="store_true")
    args = ap.parse_args()

    if args.sheet.resolve().is_relative_to(Path(__file__).resolve().parents[1]):
        print(
            f"REFUSED: {args.sheet} is inside this repo.\n"
            "  A filled workbook carries a participant's address and their own "
            "claims about\n  their competence. The repo is public. Keep it outside.",
            file=sys.stderr,
        )
        return 1

    print(f"sheet {args.sheet}")
    rows = read_sheet(args.sheet)
    print(f"  {len(rows)} non-empty row(s)")

    if args.registry_from_vault:
        try:
            reg = load(args.vault / MATRIX_SUBDIR, args.vault / DOMAIN_MAP_SUBDIR)
        except MatrixError as e:
            print(f"REFUSED:\n  {e}", file=sys.stderr)
            return 1
        known = {u.unit_key for u in reg.units}
        print(f"  registry from the vault sources: {len(known)} units")
        rest = None
    else:
        url = os.environ.get("SUPABASE_URL", "")
        key = os.environ.get("SUPABASE_SECRET_KEY", "") or os.environ.get(
            "SUPABASE_SERVICE_ROLE_KEY", ""
        )
        if not url or not key:
            print(
                "needs SUPABASE_URL and SUPABASE_SECRET_KEY, or --registry-from-vault "
                "to validate\n  against the vault sources while the project is "
                "unreachable.",
                file=sys.stderr,
            )
            return 2
        print(f"  url {url}")
        print(f"  key {key if args.show else mask(key)}")
        rest = Rest(url, key)
        known = {r["unit_key"] for r in rest.select("competency_unit", "unit_key")}
        print(f"  registry from the database: {len(known)} units")

    if len(known) != 41:
        print(
            f"\nREFUSED: the registry has {len(known)} units, not 41. Seed it first.",
            file=sys.stderr,
        )
        return 1

    try:
        clean = validate(rows, known)
    except MatrixError as e:
        print(f"\nREFUSED\n  {e}", file=sys.stderr)
        return 1

    by_status: dict[str, int] = {}
    for r in clean:
        by_status[r["claim_status"]] = by_status.get(r["claim_status"], 0) + 1
    print(f"\n{len(clean)} row(s) valid: {by_status}")
    claimed_zero = sum(1 for r in clean if r["claim_status"] == "claimed" and r["claimed_level"] == 0)
    print(f"  explicit claims of level 0: {claimed_zero}  (distinct from no-claim)")
    unmentioned = sorted(known - {r['unit_key'] for r in clean})
    if unmentioned:
        print(f"  units absent from the sheet entirely: {len(unmentioned)} ({unmentioned[:5]}…)")

    if not args.apply:
        print("\nvalidate-only, nothing written. Re-run with --apply and --email.")
        return 0

    if not args.email or not EMAIL_RE.match(args.email):
        print("\n--apply needs a valid --email for the subject.", file=sys.stderr)
        return 2
    if rest is None:
        print(
            "\n--apply cannot be combined with --registry-from-vault: there is no "
            "database to write to.",
            file=sys.stderr,
        )
        return 2

    payload = [
        {
            "subject_email": args.email.strip().lower(),
            "unit_key": r["unit_key"],
            "claimed_level": r["claimed_level"],
            "claim_status": r["claim_status"],
            "note": r["note"],
            "source_sheet": args.sheet.name,
        }
        for r in clean
    ]
    rest.insert("self_assessment_intake", payload)
    print(f"\nwrote {len(payload)} row(s) for {args.email.strip().lower()}")
    print("  self_assessment_intake is append-only and service-role only.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
