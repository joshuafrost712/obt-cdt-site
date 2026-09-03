#!/usr/bin/env python3
"""Create portal accounts with generated passwords, for a named cohort.

Why this exists rather than letting people register themselves. Several of this
cohort are on connections where a confirmation email is a coin toss, and the
portal has **no password reset flow** (memory note `honest-eval-no-password-reset`
records the same gap on the sibling app). A pre-created account with a password
Joshua hands over in the same message as the link is the path with the fewest
places to fail. The cost is that a password travels by email, which is stated in
the covering message rather than hidden.

`member_allowlist` gates account creation: `handle_new_portal_user()` raises
`insufficient_privilege` for an address that is not on it. So run
`seed_allowlist.py --apply` first, or every create here fails identically.

**No participant address is ever written into this repository.** The workbook
stays outside it and the credential file is refused if its path resolves inside
any git working tree. Addresses are masked in stdout unless you ask otherwise.

Usage
-----
    set -a; . ~/.claude/secrets/obt-cdt-supabase.env; set +a
    export PORTAL_URL="$OBT_CDT_SUPABASE_URL"
    export PORTAL_SECRET_KEY="$OBT_CDT_SUPABASE_SECRET_KEY"

    # look at what would happen (default; writes nothing, creates nothing)
    python3 scripts/create_portal_accounts.py --xlsx /path/signup.xlsx --tab 'Psalms 2026'

    # do it
    python3 scripts/create_portal_accounts.py --xlsx /path/signup.xlsx \
        --tab 'Psalms 2026' \
        --exclude someone@example.org --exclude other@example.org \
        --extra 'late.joiner@example.org=Late Joiner' \
        --out ~/Documents/obt-cdt-accounts.csv --apply

Environment
-----------
    PORTAL_URL         https://<project-ref>.supabase.co
    PORTAL_SECRET_KEY  the `sb_secret_...` key (service role). Read from the
                       environment, never argv, which is world-readable.
"""

import argparse
import csv
import json
import os
import re
import secrets
import subprocess
import sys
import urllib.error
import urllib.request

EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")

# Readable, unambiguous, and safe to dictate over a bad phone line. No words that
# sound alike, nothing that reads as an instruction, nothing that could land in a
# spam filter. Three of these plus two digits clears the project's 12-character
# minimum with room to spare.
WORDS = """
anchor amber arbor autumn basil beacon birch bridge canyon cedar cello cinder
clover cobalt comet copper coral cove crimson crystal delta dune ember fable
falcon fern forest garnet granite harbor harvest hazel heron indigo island
ivory jasper juniper kettle lagoon lantern laurel ledger linen lotus lumen
maple marble meadow mesa mica mint mosaic nectar nimbus oasis olive onyx opal
orchard otter pebble pepper pewter pine plateau prairie quartz quill ravine
reef ridge river rowan saffron sage sandal sequoia shale silver slate solstice
sparrow spruce summit tamarind teal thistle timber topaz tulip tundra umber
valley velvet verdant vessel violet walnut willow window winter zephyr zinnia
""".split()


def normalize(address: str) -> str:
    """Match `handle_new_portal_user`'s rule exactly: lower(btrim(email))."""
    return address.strip().lower()


def mask(address: str) -> str:
    local, _, domain = address.partition("@")
    if len(local) <= 2:
        return f"{local[:1] or '?'}*@{domain}"
    return f"{local[0]}{'*' * (len(local) - 2)}{local[-1]}@{domain}"


def passphrase() -> str:
    """Three words and two digits, hyphen separated.

    Chosen over a random string because these are typed by hand, often on a
    phone, often by someone reading them off another screen. A password that is
    mistyped four times is a support email, which is the thing this whole script
    exists to avoid.
    """
    words = [secrets.choice(WORDS) for _ in range(3)]
    return "-".join(words) + "-" + f"{secrets.randbelow(90) + 10}"


def refuse_if_in_git(path: str) -> None:
    """A credential file inside a git tree is a credential file that ships.

    This repository is public, and the campaign has twice caught prose leaking
    into it through a harness and through a generator (program findings 18 and
    24). A list of live passwords is the same class of mistake with a worse
    ending, so the check is a refusal and not a warning.
    """
    directory = os.path.dirname(os.path.abspath(path)) or "."
    try:
        result = subprocess.run(
            ["git", "-C", directory, "rev-parse", "--show-toplevel"],
            capture_output=True, text=True, timeout=10,
        )
    except (OSError, subprocess.SubprocessError):
        return
    if result.returncode == 0:
        sys.exit(
            f"Refusing to write credentials inside a git working tree.\n"
            f"  {os.path.abspath(path)}\n"
            f"  is inside {result.stdout.strip()}\n"
            f"Choose a path outside every repository, e.g. ~/Documents/."
        )


def harvest(path: str, tab: str) -> dict[str, str]:
    """Address -> display name, from one named tab."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl is not installed: python3 -m pip install openpyxl")

    workbook = load_workbook(path, read_only=True, data_only=True)
    if tab not in workbook.sheetnames:
        sys.exit(f"No tab named {tab!r}. Tabs are: {', '.join(workbook.sheetnames)}")

    sheet = workbook[tab]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        sys.exit(f"Tab {tab!r} is empty.")

    headers = [str(c).strip() if c is not None else "" for c in rows[0]]

    def column(*names):
        for name in names:
            if name in headers:
                return headers.index(name)
        return None

    email_at = column("Email", "Primary Email")
    if email_at is None:
        sys.exit(f"Tab {tab!r} has no Email column. Headers: {headers}")
    first_at, last_at = column("First Name"), column("Family Name", "Last Name")

    def cell(row, index):
        if index is None or index >= len(row) or row[index] is None:
            return ""
        return str(row[index]).strip()

    found: dict[str, str] = {}
    for row in rows[1:]:
        raw = cell(row, email_at)
        match = EMAIL_RE.search(raw)
        if not match:
            continue
        name = " ".join(p for p in (cell(row, first_at), cell(row, last_at)) if p)
        found[normalize(match.group(0))] = name
    return found


def auth(method: str, path: str, body=None):
    url = os.environ["PORTAL_URL"].rstrip("/") + "/auth/v1" + path
    key = os.environ["PORTAL_SECRET_KEY"]
    data = json.dumps(body).encode() if body is not None else None
    request = urllib.request.Request(
        url, data=data, method=method,
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            raw = response.read().decode()
            return json.loads(raw) if raw.strip() else {}
    except urllib.error.HTTPError as error:
        return {"__error__": f"{error.code}: {error.read().decode()[:300]}"}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", required=True, help="exported workbook, kept outside this repo")
    parser.add_argument("--tab", required=True, help="which tab names the cohort, e.g. 'Psalms 2026'")
    parser.add_argument("--exclude", action="append", default=[], metavar="EMAIL",
                        help="address to leave out; repeatable. Use for withdrawals.")
    parser.add_argument("--extra", action="append", default=[], metavar="EMAIL=NAME",
                        help="address to add that the tab misses; repeatable.")
    parser.add_argument("--out", help="where to write the credential CSV (required with --apply)")
    parser.add_argument("--apply", action="store_true", help="actually create; omit for a dry run")
    parser.add_argument("--show", action="store_true", help="print full addresses instead of masking")
    args = parser.parse_args()

    people = harvest(args.xlsx, args.tab)
    print(f"{len(people)} addresses on tab {args.tab!r}")

    for spec in args.extra:
        address, _, name = spec.partition("=")
        people[normalize(address)] = name.strip()
        print(f"  + added by hand: {mask(normalize(address))}  {name.strip()}")

    for address in args.exclude:
        key = normalize(address)
        if people.pop(key, None) is None:
            print(f"  !! --exclude {mask(key)} matched nobody on this tab", file=sys.stderr)
        else:
            print(f"  - excluded: {mask(key)}")

    print(f"\n{len(people)} accounts to create")

    if args.apply:
        if not args.out:
            sys.exit("--apply needs --out, or the passwords are generated and lost.")
        refuse_if_in_git(args.out)
        for name in ("PORTAL_URL", "PORTAL_SECRET_KEY"):
            if not os.environ.get(name):
                sys.exit(f"Set {name} (see the module docstring).")

    existing = {}
    if args.apply:
        page = auth("GET", "/admin/users?per_page=1000")
        if "__error__" in page:
            sys.exit(f"Could not list existing users: {page['__error__']}")
        existing = {normalize(u["email"]): u for u in page.get("users", []) if u.get("email")}
        print(f"{len(existing)} accounts already exist on the project")

    show = (lambda a: a) if args.show else mask
    results = []

    for address in sorted(people):
        name = people[address]
        if not args.apply:
            results.append((address, name, "would-create", ""))
            continue
        if address in existing:
            print(f"  = {show(address):<38} already has an account, left alone")
            results.append((address, name, "already-existed", ""))
            continue
        password = passphrase()
        created = auth("POST", "/admin/users", {
            "email": address,
            "password": password,
            "email_confirm": True,
            "user_metadata": {"full_name": name} if name else {},
        })
        if "__error__" in created:
            print(f"  ! {show(address):<38} FAILED  {created['__error__']}", file=sys.stderr)
            results.append((address, name, "failed", ""))
            continue
        print(f"  + {show(address):<38} created")
        results.append((address, name, "created", password))

    created = [r for r in results if r[2] == "created"]
    failed = [r for r in results if r[2] == "failed"]
    kept = [r for r in results if r[2] == "already-existed"]

    if not args.apply:
        print("\nDry run. Would create:")
        for address, name, _, _ in results:
            print(f"  {show(address):<38} {name}")
        print("\nNothing was created and no password was generated. Re-run with --apply --out PATH.")
        return

    with open(args.out, "w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["name", "email", "password", "status"])
        for address, name, status, password in results:
            writer.writerow([name, address, password, status])
    os.chmod(args.out, 0o600)

    print(f"\ncreated {len(created)}, already existed {len(kept)}, failed {len(failed)}")
    print(f"credentials written to {args.out} (mode 600)")
    if failed:
        print("\nEvery failure above is usually the same thing: the address is not on")
        print("`member_allowlist`. Run seed_allowlist.py --apply and try again.")


if __name__ == "__main__":
    main()
